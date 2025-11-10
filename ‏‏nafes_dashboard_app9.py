# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from io import BytesIO
from PIL import Image as PILImage
from PIL import ImageOps
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# دعم العربية في الرسوم
import arabic_reshaper
from bidi.algorithm import get_display

st.set_page_config(page_title="لوحة تحليل بيانات المدارس - نافس (نهائية + تصدير شامل)", layout="wide")

st.title("الداشبورد الشامل في تحليل بيانات نافس لمدارس تعليم الطائف")
st.caption("تحليل نسب الأداء، التحسّن/الانخفاض، ترتيب المدرسة (إدارة/مملكة)، معالجة الأعمدة المكرّرة، واختيار وضع العرض الزمني، مع تنزيل كل النتائج في ملف .")

@st.cache_data
def load_excel(uploaded_file_or_path):
    if uploaded_file_or_path is None:
        return None
    try:
        if isinstance(uploaded_file_or_path, str):
            xls = pd.ExcelFile(uploaded_file_or_path)
            df = pd.read_excel(uploaded_file_or_path, sheet_name=0)
        else:
            xls = pd.ExcelFile(uploaded_file_or_path)
            df = pd.read_excel(uploaded_file_or_path, sheet_name=0)

        # توحيد أسماء الأعمدة
        df.columns = [str(c).strip() for c in df.columns]

        # معالجة الأعمدة المكررة: دمج أول قيمة غير فارغة ثم حذف الأعمدة الزائدة
        dup_names = pd.Index(df.columns)[pd.Index(df.columns).duplicated()].unique().tolist()
        dedup_report = {}
        if dup_names:
            for name in dup_names:
                same_cols = [c for c in df.columns if c == name]
                merged = df[same_cols].bfill(axis=1).iloc[:, 0]
                df[name] = merged
                for extra in same_cols[1:]:
                    df.drop(columns=[extra], inplace=True)
                dedup_report[name] = same_cols

        return df, xls.sheet_names, dedup_report
    except Exception as e:
        st.error(f"خطأ أثناء قراءة الملف: {e}")
        return None

def guess_col(df, candidates):
    for c in df.columns:
        name = str(c).replace("ـ", "").replace("_", "").replace("  ", " ").strip()
        for cand in candidates:
            if cand in name:
                return c
    return None

def safe_numeric(s):
    return pd.to_numeric(s, errors="coerce")

def extract_year_series(series):
    def parse_one(x):
        if pd.isna(x):
            return np.nan
        xs = str(x)
        import re
        m = re.search(r"(20\d{2}|19\d{2}|14\d{2})", xs)
        if m:
            return int(m.group(1))
        try:
            return int(float(xs))
        except Exception:
            return np.nan
    return series.apply(parse_one)

def normalize_pct_row(row, cols, tol=2.0):
    s = row[cols].sum(skipna=True)
    if np.isfinite(s) and (100 - tol) <= s <= (100 + tol) and s != 0:
        row[cols] = row[cols] * (100.0 / s)
    return row

def make_unique_names(cols):
    cols = list(cols)
    new_cols = []
    seen = {}
    for c in cols:
        if c not in seen:
            seen[c] = 0
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")
    return new_cols

def dfs_to_excel_bytes(sheets_dict, images=None, charts_sheet_name="رسوم_بيانية"):
    """
    sheets_dict: dict[str, pandas.DataFrame]
    images: list[tuple[str, PIL.Image.Image]]   # [(title, pil_image), ...]
    charts_sheet_name: str
    """
    import pandas as pd
    from io import BytesIO
    import tempfile, os
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    # 1) اكتب الجداول أولًا في ملف مؤقت
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    def _clean_sheet_name(name: str) -> str:
        bad = '[]:*?/\\'
        safe = ''.join(c for c in str(name) if c not in bad)[:31]
        return safe if safe else "Sheet"

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            try:
                df.to_excel(writer, sheet_name=_clean_sheet_name(sheet_name), index=False)
            except Exception:
                df.reset_index().to_excel(writer, sheet_name=_clean_sheet_name(sheet_name), index=False)

    # 2) أضف الرسوم (إن وُجدت) إلى ورقة منفصلة
    if images:
        wb = load_workbook(tmp_path)
        sheet_name = _clean_sheet_name(charts_sheet_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        row = 1
        temp_img_paths = []

        try:
            for title, pil_img in images:
                # عنوان الرسم
                ws.cell(row=row, column=1, value=str(title))
                row += 1

                # خزّن الصورة مؤقتًا كملف PNG ثم أدرجها
                img_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                temp_img_paths.append(img_tmp.name)
                pil_img.save(img_tmp, format="PNG")
                img_tmp.close()

                xl_img = XLImage(img_tmp.name)
                xl_img.anchor = f"A{row}"
                ws.add_image(xl_img)

                row += 30  # مسافة بين الرسوم

            wb.save(tmp_path)
        finally:
            for p in temp_img_paths:
                try:
                    os.remove(p)
                except Exception:
                    pass

    with open(tmp_path, "rb") as f:
        data = f.read()

    try:
        os.remove(tmp_path)
    except Exception:
        pass

    return data

def fig_to_pil(fig):
    """حوّل شكل matplotlib إلى صورة PIL."""
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    img = PILImage.open(buf)
    img = ImageOps.expand(img, border=0, fill="white")
    return img

def ar(text):
    """
    تهيئة النص العربي للرسم: ربط الحروف + اتجاه من اليمين لليسار.
    تُستخدم فقط داخل matplotlib، وليس في جداول/واجهة Streamlit.
    """
    if text is None:
        return ""
    reshaped = arabic_reshaper.reshape(str(text))
    return get_display(reshaped)

# ----- تحميل الملف -----
default_path = "/mnt/data/نسخة من تقرير_نافس_20_08_2025.xlsx"
use_default = st.toggle("استخدام ملف المثال المرفق", value=True, help="يمكنك إلغاء التفعيل لتحميل ملف آخر.")
uploaded = default_path if use_default else st.file_uploader("ارفع ملف Excel (xlsx)", type=["xlsx"])

loaded = load_excel(uploaded)
if not loaded:
    st.stop()

df, sheet_names, dedup_report = loaded

# تقرير الأعمدة المدموجة
if dedup_report:
    with st.expander("تم اكتشاف أعمدة مكررة وتم دمجها تلقائيًا (اضغط للاطلاع)"):
        for name, cols in dedup_report.items():
            st.write(f"**{name}** ← تم دمج الأعمدة: {', '.join(cols)}")

# ----- تخمين أعمدة -----
school_col   = guess_col(df, ["اسم المدرسة", "المدرسة"])
year_col     = guess_col(df, ["العام الدراسي", "العام", "سنة"])
admin_col    = guess_col(df, ["إدارة التعليم"])
domain_col   = guess_col(df, ["المجال"])
subdom_col   = guess_col(df, ["المجال الفرعي", "فرعي"])
score_col    = guess_col(df, ["المتوسط"])
king_avg_col = guess_col(df, ["المتوسط على مستوى المملكة"])
admin_avg_col= guess_col(df, ["المتوسط على مستوى إدارة التعليم"])

pct_very_low = guess_col(df, ["نسبة الطلاب في مستوى الأداء المنخفض جدا"])
pct_low      = guess_col(df, ["نسبة الطلاب في مستوى الأداء المنخفض"])
pct_mid      = guess_col(df, ["نسبة الطلاب في مستوى الأداء المتوسط"])
pct_high     = guess_col(df, ["نسبة الطلاب في مستوى الأداء المرتفع"])

count_col    = guess_col(df, ["مجموع طلاب الصف", "مجموع المختبرين", "عدد الطلاب", "طلاب"])
grade_col    = guess_col(df, ["الصف", "صف"])
gender_col   = guess_col(df, ["جنس المدرسة", "الجنس", "بنين", "بنات"])
type_col     = guess_col(df, ["نوع المدرسة", "نوع", "حكومي", "أهلي"])

with st.expander("تأكيد/تعديل أعمدة البيانات"):
    c1, c2, c3 = st.columns(3)
    with c1:
        school_col = st.selectbox("عمود المدرسة", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(school_col) if school_col in df.columns else 0)
        year_col   = st.selectbox("عمود العام الدراسي", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(year_col) if year_col in df.columns else 0)
        score_col  = st.selectbox("عمود متوسط أداء المدرسة", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(score_col) if score_col in df.columns else 0)
    with c2:
        domain_col = st.selectbox("عمود المجال", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(domain_col) if domain_col in df.columns else 0)
        subdom_col = st.selectbox("عمود المجال الفرعي", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(subdom_col) if subdom_col in df.columns else 0)
        admin_col  = st.selectbox("عمود إدارة التعليم", [None] + list(df.columns),
                                  index=([None] + list(df.columns)).index(admin_col) if admin_col in df.columns else 0)
    with c3:
        king_avg_col  = st.selectbox("عمود المتوسط على مستوى المملكة", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(king_avg_col) if king_avg_col in df.columns else 0)
        admin_avg_col = st.selectbox("عمود المتوسط على مستوى إدارة التعليم", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(admin_avg_col) if admin_avg_col in df.columns else 0)
        count_col     = st.selectbox("عمود عدد الطلاب (اختياري)", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(count_col) if count_col in df.columns else 0)
        grade_col     = st.selectbox("عمود الصف (اختياري)", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(grade_col) if grade_col in df.columns else 0)
        gender_col    = st.selectbox("عمود جنس المدرسة (اختياري)", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(gender_col) if gender_col in df.columns else 0)
        type_col      = st.selectbox("عمود نوع المدرسة (اختياري)", [None] + list(df.columns),
                                     index=([None] + list(df.columns)).index(type_col) if type_col in df.columns else 0)

    st.markdown("**أعمدة النسب (مجموعها ≈ 100%):**")
    c4, c5, c6, c7 = st.columns(4)
    with c4:
        pct_very_low = st.selectbox("منخفض جدًا %", [None] + list(df.columns),
                                    index=([None] + list(df.columns)).index(pct_very_low) if pct_very_low in df.columns else 0)
    with c5:
        pct_low      = st.selectbox("منخفض %", [None] + list(df.columns),
                                    index=([None] + list(df.columns)).index(pct_low) if pct_low in df.columns else 0)
    with c6:
        pct_mid      = st.selectbox("متوسط %", [None] + list(df.columns),
                                    index=([None] + list(df.columns)).index(pct_mid) if pct_mid in df.columns else 0)
    with c7:
        pct_high     = st.selectbox("مرتفع %", [None] + list(df.columns),
                                    index=([None] + list(df.columns)).index(pct_high) if pct_high in df.columns else 0)

# تحويلات رقمية (بدون الجنس ونوع المدرسة)
for c in [score_col, king_avg_col, admin_avg_col, pct_very_low, pct_low, pct_mid, pct_high, count_col, grade_col]:
    if c:
        df[c] = safe_numeric(df[c])

# تجهيز عام/سنة رقمية
if year_col:
    df["__year"] = extract_year_series(df[year_col])
else:
    df["__year"] = np.nan

# تطبيع تقريبي للنسب إن لزم
pct_cols = [c for c in [pct_very_low, pct_low, pct_mid, pct_high] if c]
if pct_cols:
    df = df.apply(lambda r: normalize_pct_row(r, pct_cols), axis=1)

# قائمة لحفظ كل الرسوم (لوحة عامة + مدرسة + تنبؤ) للتصدير إلى الإكسل
chart_images = []

# ====== لوحة عامة ======
st.subheader("لوحة عامة (جميع المدارس) – النسب حسب المجال/الفرعي")

filter_year = None
filter_grade = None
filter_gender = None
filter_type = None
filter_domain = None
filter_subdom = None

colY, colG, colSex, colType, colA, colB = st.columns(6)

# اختيار العام
years_available_all = []
if "__year" in df.columns:
    years_tmp = df["__year"].dropna().unique().tolist()
    years_available_all = sorted({int(y) for y in years_tmp if pd.notna(y)})
with colY:
    if years_available_all:
        year_options = ["جميع الأعوام"] + years_available_all
        filter_year = st.selectbox("اختر العام", year_options, index=0)
    else:
        filter_year = "جميع الأعوام"
        st.write("لا يوجد عام دراسي محدد في البيانات.")

# اختيار الصف
with colG:
    if grade_col:
        grade_options = ["جميعها", 3, 6, 9]
        filter_grade = st.selectbox("اختر الصف", grade_options, index=0)
    else:
        filter_grade = "جميعها"

# اختيار جنس المدرسة
with colSex:
    if gender_col:
        gender_options = ["الكل", "بنين", "بنات"]
        filter_gender = st.selectbox("جنس المدرسة", gender_options, index=0)
    else:
        filter_gender = "الكل"

# اختيار نوع المدرسة
with colType:
    if type_col:
        type_options = ["الكل", "حكومي", "أهلي"]
        filter_type = st.selectbox("نوع المدرسة", type_options, index=0)
    else:
        filter_type = "الكل"

# اختيار المجال
with colA:
    if domain_col:
        opts = ["(الكل)"] + sorted(df[domain_col].dropna().astype(str).unique().tolist())
        filter_domain = st.selectbox("تصفية المجال", opts)
# اختيار المجال الفرعي
with colB:
    if subdom_col:
        opts2 = ["(الكل)"] + sorted(df[subdom_col].dropna().astype(str).unique().tolist())
        filter_subdom = st.selectbox("تصفية المجال الفرعي", opts2)

# تطبيق الفلاتر
dff = df.copy()
if filter_year and filter_year != "جميع الأعوام" and "__year" in dff.columns:
    dff = dff[dff["__year"] == filter_year]

if grade_col and filter_grade and filter_grade != "جميعها":
    try:
        grade_value = int(filter_grade)
        dff = dff[dff[grade_col] == grade_value]
    except Exception:
        pass

if gender_col and filter_gender and filter_gender != "الكل":
    dff = dff[dff[gender_col].astype(str) == filter_gender]

if type_col and filter_type and filter_type != "الكل":
    dff = dff[dff[type_col].astype(str) == filter_type]

if domain_col and filter_domain and filter_domain != "(الكل)":
    dff = dff[dff[domain_col] == filter_domain]
if subdom_col and filter_subdom and filter_subdom != "(الكل)":
    dff = dff[dff[subdom_col] == filter_subdom]

general_sheets = {}
if pct_cols and len(pct_cols) == 4:
    st.write("**متوسط النسب عبر المدارس**")
    grp_cols = []
    if domain_col:
        grp_cols.append(domain_col)
    if subdom_col:
        grp_cols.append(subdom_col)

    if grp_cols:
        # متوسط النِّسَب حسب (المجال / المجال الفرعي)
        agg_raw = dff.groupby(grp_cols, dropna=True)[pct_cols].mean()
        agg = agg_raw.sort_values(by=pct_high, ascending=False).copy()
        agg.columns = make_unique_names(agg.columns)
        t = agg.reset_index()
        st.dataframe(t)
        general_sheets["متوسط_النسب_لوحة_عامة"] = t

        # 🎨 رسم بياني مكدّس: المحور الأفقي = المجال الفرعي (وتحته المجال)، العمودي = النِّسَب
        if domain_col and subdom_col:
            plot_df = agg_raw.reset_index().sort_values([domain_col, subdom_col])
            if not plot_df.empty:
                fig, ax = plt.subplots(figsize=(12, 5))

                x = np.arange(len(plot_df))
                bottoms = np.zeros(len(plot_df))

                # أسماء المستويات (عاديّة للواجهة، سنحوّلها بـ ar() داخل الرسم فقط)
                nice_names = {
                    pct_very_low: "منخفض جدًا",
                    pct_low:      "منخفض",
                    pct_mid:      "متوسط",
                    pct_high:     "مرتفع",
                }

                # ألوان ثابتة لمستويات الأداء
                colors_map = {
                    pct_very_low: "#1f77b4",  # أزرق
                    pct_low:      "#ff7f0e",  # برتقالي
                    pct_mid:      "#2ca02c",  # أخضر
                    pct_high:     "#d62728",  # أحمر
                }

                # أسماء الألوان بالعربي
                colors_names = {
                    pct_very_low: "أزرق",
                    pct_low:      "برتقالي",
                    pct_mid:      "أخضر",
                    pct_high:     "أحمر",
                }

                legend_rows = []

                for col in pct_cols:
                    vals = plot_df[col].values
                    base_label = nice_names.get(col, str(col))  # نص عربي عادي
                    plot_label = ar(base_label)                 # نسخة مجهّزة للرسم فقط
                    color = colors_map.get(col, None)

                    ax.bar(x, vals, bottom=bottoms, label=plot_label, color=color)
                    bottoms += vals

                    legend_rows.append({
                        "مستوى الأداء": base_label,          # عربي عادي، بدون ar()
                        "اسم اللون": colors_names.get(col, ""),
                        "كود اللون": color,
                    })

                labels = [
                    ar(f"{row[subdom_col]}\n({row[domain_col]})")
                    for _, row in plot_df.iterrows()
                ]
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha="right")

                ax.set_ylabel(ar("النسبة %"))
                ax.set_title(ar("توزيع مستويات الأداء حسب المجال الفرعي (مرتّبة حسب المجال)"))

                # ❌ لا نعرض Legend داخل الرسم
                fig.tight_layout()
                st.pyplot(fig)

                # ✅ جدول ألوان مستويات الأداء تحت الرسم، بالعربي الصحيح وباسم اللون
                if legend_rows:
                    legend_df = pd.DataFrame(legend_rows).drop_duplicates()
                    st.write("**جدول ألوان مستويات الأداء (بديل لمربع المعلومات في الرسم)**")
                    st.dataframe(legend_df)

                # إضافة الرسم إلى قائمة الرسوم للتصدير في الإكسل
                chart_images.append(
                    (ar("توزيع المستويات حسب المجال الفرعي - لوحة عامة"), fig_to_pil(fig))
                )

    else:
        one = dff[pct_cols].mean().to_frame("متوسط %").reset_index().rename(columns={"index": "النسبة"})
        st.dataframe(one)
        general_sheets["متوسط_النسب_لوحة_عامة"] = one

# ====== تحليل مدرسة محددة ======
st.markdown("---")
st.subheader("تحليل مدرسة محددة عبر الأعوام مع النسب والترتيب")

school_sel = None
domain_sel = None
grade_sel_school = None
gender_sel_school = None
type_sel_school = None

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    if gender_col:
        gender_options = ["الكل", "بنين", "بنات"]
        gender_sel_school = st.selectbox("جنس المدرسة (اختياري)", gender_options, index=0)
    else:
        gender_sel_school = "الكل"

with col2:
    if type_col:
        type_options = ["الكل", "حكومي", "أهلي"]
        type_sel_school = st.selectbox("نوع المدرسة (اختياري)", type_options, index=0)
    else:
        type_sel_school = "الكل"

with col3:
    if school_col:
        df_schools = df.copy()
        if gender_col and gender_sel_school and gender_sel_school != "الكل":
            df_schools = df_schools[df_schools[gender_col].astype(str) == gender_sel_school]
        if type_col and type_sel_school and type_sel_school != "الكل":
            df_schools = df_schools[df_schools[type_col].astype(str) == type_sel_school]
        schools = sorted(df_schools[school_col].dropna().astype(str).unique().tolist())
        school_sel = st.selectbox("اختر المدرسة", schools if schools else ["لا توجد بيانات"])

with col4:
    if domain_col:
        dom_opts = ["(الكل)"] + sorted([x for x in df[domain_col].dropna().unique().tolist() if str(x).strip()])
        domain_sel = st.selectbox("تصفية بالمجال (اختياري)", dom_opts)

with col5:
    if grade_col:
        grade_options = ["جميعها", 3, 6, 9]
        grade_sel_school = st.selectbox("تصفية بالصف (اختياري)", grade_options, index=0)

export_sheets = {}

if school_sel:
    sdf_all = df.copy()
    if gender_col and gender_sel_school and gender_sel_school != "الكل":
        sdf_all = sdf_all[sdf_all[gender_col].astype(str) == gender_sel_school]
    if type_col and type_sel_school and type_sel_school != "الكل":
        sdf_all = sdf_all[sdf_all[type_col].astype(str) == type_sel_school]

    sdf_all = sdf_all[sdf_all[school_col] == school_sel]

    if domain_col and domain_sel and domain_sel != "(الكل)":
        sdf_all = sdf_all[sdf_all[domain_col] == domain_sel]
    if grade_col and grade_sel_school and grade_sel_school != "جميعها":
        try:
            gv = int(grade_sel_school)
            sdf_all = sdf_all[sdf_all[grade_col] == gv]
        except Exception:
            pass

    years_available = sorted(sdf_all["__year"].dropna().unique().tolist()) if "__year" in sdf_all.columns else []
    latest_year = years_available[-1] if years_available else None

    st.markdown("##### الوضع الزمني")
    mode = st.radio("طريقة العرض", ["أحدث عام تلقائيًا", "عام محدد", "كل الأعوام"], horizontal=True)

    manual_year = None
    if mode == "عام محدد" and years_available:
        manual_year = st.selectbox("اختر العام الدراسي", years_available, index=len(years_available) - 1)
    active_year = manual_year if (mode == "عام محدد" and manual_year is not None) else latest_year

    # اتجاه الأداء العام للمدرسة عبر الأعوام
    if score_col:
        trend = sdf_all.sort_values("__year").groupby("__year")[score_col].mean()
        st.line_chart(trend)
        if not trend.empty:
            fig = plt.figure()
            ax = fig.gca()
            trend.plot(ax=ax)
            ax.set_xlabel(ar("العام"))
            ax.set_ylabel(ar("متوسط المدرسة"))
            ax.set_title(ar(f"اتجاه متوسط الأداء - {school_sel}"))
            chart_images.append(("اتجاه متوسط الأداء - المدرسة", fig_to_pil(fig)))

    # كل الأعوام
    if mode == "كل الأعوام":
        if pct_cols:
            yoy_tbl = (sdf_all.groupby("__year")[pct_cols].mean().sort_index())
            yoy_tbl.columns = make_unique_names(yoy_tbl.columns)
            t1 = yoy_tbl.reset_index().rename(columns={"__year": "العام"})
            st.write("**توزيع النسب عبر جميع الأعوام (متوسط لكل عام)**")
            st.dataframe(t1)
            export_sheets["نسب_كل_الأعوام"] = t1

            st.write("**فروق سنة-سنة (YoY Δ) لكل نسبة**")
            yoy_delta = yoy_tbl.diff().dropna()
            yoy_delta.columns = [f"Δ {c}" for c in yoy_delta.columns]
            t2 = yoy_delta.reset_index().rename(columns={"__year": "العام"})
            st.dataframe(t2)
            export_sheets["فروق_YoY"] = t2

            if pct_high in yoy_tbl.columns:
                fig = plt.figure()
                ax = fig.gca()
                yoy_tbl[pct_high].plot(ax=ax)
                ax.set_xlabel(ar("العام"))
                ax.set_ylabel(ar(pct_high))
                ax.set_title(ar(f"{pct_high} عبر الأعوام - {school_sel}"))
                chart_images.append((f"{pct_high} عبر الأعوام - المدرسة", fig_to_pil(fig)))

        if score_col:
            by_school_year = df.groupby([school_col, "__year"])[score_col].mean().reset_index(name="متوسط_المدرسة")
            rank_admin_list, rank_nat_list = [], []
            for y in years_available:
                row_this = by_school_year[(by_school_year[school_col] == school_sel) & (by_school_year["__year"] == y)]
                my_rank_admin = total_admin = None
                if admin_col and not df.loc[df[school_col] == school_sel, admin_col].empty:
                    my_admin = df.loc[df[school_col] == school_sel, admin_col].iloc[0]
                    same_admin = df[df[admin_col] == my_admin]
                    by_sch_admin = same_admin.groupby([school_col, "__year"])[score_col].mean().reset_index(name="متوسط_المدرسة")
                    in_year = by_sch_admin[by_sch_admin["__year"] == y].copy()
                    if not in_year.empty and not row_this.empty:
                        in_year["rank_admin"] = in_year["متوسط_المدرسة"].rank(ascending=False, method="min")
                        my_rank_admin = int(in_year[in_year[school_col] == school_sel]["rank_admin"].iloc[0])
                        total_admin = int(in_year.shape[0])
                rank_admin_list.append((y, my_rank_admin, total_admin))

                my_rank_nat = total_nat = None
                all_year = by_school_year[by_school_year["__year"] == y].copy()
                if not all_year.empty and not row_this.empty:
                    all_year["rank_nat"] = all_year["متوسط_المدرسة"].rank(ascending=False, method="min")
                    my_rank_nat = int(all_year[all_year[school_col] == school_sel]["rank_nat"].iloc[0])
                    total_nat = int(all_year.shape[0])
                rank_nat_list.append((y, my_rank_nat, total_nat))

            rank_admin_df = pd.DataFrame(rank_admin_list, columns=["العام", "ترتيب داخل الإدارة", "عدد مدارس الإدارة"])
            rank_nat_df   = pd.DataFrame(rank_nat_list,   columns=["العام", "ترتيب على مستوى المملكة", "عدد مدارس المملكة"])
            st.write("**مسار الترتيب داخل الإدارة عبر الأعوام**")
            st.dataframe(rank_admin_df)
            st.write("**مسار الترتيب على مستوى المملكة عبر الأعوام**")
            st.dataframe(rank_nat_df)

            export_sheets["ترتيب_الإدارة_عبر_الأعوام"] = rank_admin_df
            export_sheets["ترتيب_المملكة_عبر_الأعوام"] = rank_nat_df

    # لقطة لعام واحد
    if mode in ["أحدث عام تلقائيًا", "عام محدد"] and active_year is not None:
        focus = sdf_all[sdf_all["__year"] == active_year].copy()

        cols_show = []
        if domain_col:
            cols_show.append(domain_col)
        if subdom_col:
            cols_show.append(subdom_col)
        cols_show += [c for c in pct_cols if c]

        if not focus.empty and cols_show:
            st.write(f"**توزيع النسب للعام {active_year}** (يجب أن تقارب 100% لكل صف)")
            cols_unique = []
            seen_cols = set()
            for c in cols_show:
                if c not in seen_cols:
                    cols_unique.append(c)
                    seen_cols.add(c)
            focus_view = focus[cols_unique].copy()
            focus_view.columns = make_unique_names(focus_view.columns)
            t3 = focus_view.reset_index(drop=True)
            st.dataframe(t3)
            export_sheets[f"نسب_{active_year}"] = t3

            if len([c for c in pct_cols if c]) == 4:
                pct_means = focus[pct_cols].mean()
                fig = plt.figure()
                ax = fig.gca()
                pct_means.plot(kind="bar", ax=ax)
                ax.set_title(ar(f"متوسط توزيع مستويات الأداء - {active_year}"))
                chart_images.append((f"توزيع المستويات {active_year} - المدرسة", fig_to_pil(fig)))

        prev_year = active_year - 1
        prev = sdf_all[sdf_all["__year"] == prev_year].copy()
        if not focus.empty and not prev.empty and pct_cols:
            key_cols = []
            if domain_col:
                key_cols.append(domain_col)
            if subdom_col:
                key_cols.append(subdom_col)
            if not key_cols:
                cur_avg = focus[pct_cols].mean()
                prev_avg = prev[pct_cols].mean()
                delta = (cur_avg - prev_avg).rename("الفرق%")
                comp = pd.concat(
                    [cur_avg.rename("العام الحالي %"),
                     prev_avg.rename("العام السابق %"),
                     delta],
                    axis=1
                ).reset_index().rename(columns={"index": "الفئة"})
                st.write("**التغيّر عن العام السابق (بدون تفصيل تصنيفي)**")
                st.dataframe(comp)
                export_sheets[f"فروق_{prev_year}_إلى_{active_year}"] = comp
            else:
                cur_g = focus.groupby(key_cols)[pct_cols].mean()
                prev_g = prev.groupby(key_cols)[pct_cols].mean()
                mix = cur_g.join(prev_g, lsuffix="_حالي", rsuffix="_سابق", how="outer").fillna(0.0)
                for pc in pct_cols:
                    mix[f"الفرق بين العام السابق والحالي في {pc}"] = mix[f"{pc}_حالي"] - mix[f"{pc}_سابق"]
                st.write("**التغيّر عن العام السابق حسب (المجال/الفرعي)**")
                mix_reset = mix.reset_index()
                st.dataframe(mix_reset)
                export_sheets[f"فروق_{prev_year}_إلى_{active_year}"] = mix_reset

        if score_col:
            by_school_year = df.groupby([school_col, "__year"])[score_col].mean().reset_index(name="متوسط_المدرسة")
            row_this = by_school_year[(by_school_year[school_col] == school_sel) & (by_school_year["__year"] == active_year)]

            my_rank_admin = total_admin = None
            if admin_col and not df.loc[df[school_col] == school_sel, admin_col].empty:
                my_admin = df.loc[df[school_col] == school_sel, admin_col].iloc[0]
                same_admin = df[df[admin_col] == my_admin]
                by_sch_admin = same_admin.groupby([school_col, "__year"])[score_col].mean().reset_index(
                    name="متوسط_المدرسة"
                )
                in_year = by_sch_admin[by_sch_admin["__year"] == active_year].copy()
                if not in_year.empty and not row_this.empty:
                    in_year["الترتيب_داخل_الإدارة"] = in_year["متوسط_المدرسة"].rank(
                        ascending=False, method="min"
                    )
                    in_year = in_year.sort_values("الترتيب_داخل_الإدارة")
                    my_rank_admin = int(in_year[in_year[school_col] == school_sel]["الترتيب_داخل_الإدارة"].iloc[0])
                    total_admin = int(in_year.shape[0])

            my_rank_nat = total_nat = None
            all_year = by_school_year[by_school_year["__year"] == active_year].copy()
            if not all_year.empty and not row_this.empty:
                all_year["الترتيب_على_مستوى_المملكة"] = all_year["متوسط_المدرسة"].rank(
                    ascending=False, method="min"
                )
                all_year = all_year.sort_values("الترتيب_على_مستوى_المملكة")
                my_rank_nat = int(all_year[all_year[school_col] == school_sel]["الترتيب_على_مستوى_المملكة"].iloc[0])
                total_nat = int(all_year.shape[0])

            admin_avg_val = df[df["__year"] == active_year][admin_avg_col].mean(skipna=True) if admin_avg_col else None
            king_avg_val  = df[df["__year"] == active_year][king_avg_col].mean(skipna=True) if king_avg_col else None
            school_avg_val = float(row_this["متوسط_المدرسة"].iloc[0]) if not row_this.empty else None

            kpi = pd.DataFrame({
                "مؤشر": ["متوسط المدرسة", "ترتيب داخل الإدارة", "ترتيب على مستوى المملكة",
                         "متوسط الإدارة (مرجعي)", "متوسط المملكة (مرجعي)"],
                "قيمة": [f"{school_avg_val:.2f}" if school_avg_val is not None else "—",
                         f"{my_rank_admin}/{total_admin}" if my_rank_admin is not None else "—",
                         f"{my_rank_nat}/{total_nat}" if my_rank_nat is not None else "—",
                         f"{admin_avg_val:.2f}" if admin_avg_val is not None else "—",
                         f"{king_avg_val:.2f}" if king_avg_val is not None else "—"]
            })
            st.write("**مؤشرات العام المختار**")
            st.dataframe(kpi)
            export_sheets[f"KPI_{active_year}"] = kpi

# ====== تنبؤات ======
st.markdown("---")
st.subheader("التنبؤ للأعوام القادمة")

years_ahead = st.slider("عدد الأعوام المتنبأ بها", 1, 5, 3)

def forecast_linear(years, values):
    years = np.array(years, dtype=float)
    values = np.array(values, dtype=float)
    mask = ~np.isnan(years) & ~np.isnan(values)
    years = years[mask]
    values = values[mask]
    if len(np.unique(years)) < 2 or len(values) < 2:
        return None
    lr = LinearRegression().fit(years.reshape(-1, 1), values)
    future_years = np.arange(int(np.nanmax(years)) + 1, int(np.nanmax(years)) + 1 + years_ahead)
    preds = lr.predict(future_years.reshape(-1, 1))
    return pd.DataFrame({"العام": future_years, "توقع": preds})

colX, colY = st.columns(2)
with colX:
    if 'school_sel' in locals() and school_sel and school_col and year_col and score_col:
        by_school_year = df.groupby([school_col, "__year"])[score_col].mean().reset_index()
        st.write("**توقع مستوى الأداء (متوسط المدرسة)**")
        g = by_school_year[by_school_year[school_col] == school_sel].dropna()
        fc_perf = forecast_linear(g["__year"], g[score_col])
        if fc_perf is not None and not fc_perf.empty:
            st.dataframe(fc_perf)
            st.line_chart(fc_perf.set_index("العام")["توقع"])
            fig = plt.figure()
            ax = fig.gca()
            ax.plot(fc_perf["العام"].values, fc_perf["توقع"].values)
            ax.set_xlabel(ar("العام"))
            ax.set_ylabel(ar("توقع الأداء"))
            ax.set_title(ar(f"توقع الأداء - {school_sel}"))
            chart_images.append((f"توقع الأداء - {school_sel}", fig_to_pil(fig)))
            export_sheets["توقع_الأداء"] = fc_perf.reset_index(drop=True)
        else:
            st.info("بيانات غير كافية لتوقع الأداء للمدرسة المحددة.")
with colY:
    if 'school_sel' in locals() and school_sel and school_col and year_col and count_col:
        by_school_year_cnt = df.groupby([school_col, "__year"])[count_col].sum().reset_index()
        st.write("**توقع عدد الطلاب**")
        g = by_school_year_cnt[by_school_year_cnt[school_col] == school_sel].dropna()
        fc_cnt = forecast_linear(g["__year"], g[count_col])
        if fc_cnt is not None and not fc_cnt.empty:
            st.dataframe(fc_cnt)
            st.bar_chart(fc_cnt.set_index("العام")["توقع"])
            fig = plt.figure()
            ax = fig.gca()
            ax.bar(fc_cnt["العام"].values, fc_cnt["توقع"].values)
            ax.set_xlabel(ar("العام"))
            ax.set_ylabel(ar("توقع عدد الطلاب"))
            ax.set_title(ar(f"توقع عدد الطلاب - {school_sel}"))
            chart_images.append((f"توقع عدد الطلاب - {school_sel}", fig_to_pil(fig)))
            export_sheets["توقع_عدد_الطلاب"] = fc_cnt.reset_index(drop=True)
        else:
            st.info("بيانات غير كافية لتوقع عدد الطلاب للمدرسة المحددة.")

st.markdown("### تنزيل")
if general_sheets:
    general_xlsx = dfs_to_excel_bytes(general_sheets)
    st.download_button("تنزيل نتائج اللوحة العامة (Excel)", data=general_xlsx,
                       file_name="general_dashboard.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if export_sheets:
    xlsx_bytes = dfs_to_excel_bytes(export_sheets)
    default_name = f"school_results_{school_sel}.xlsx".replace(" ", "_") if 'school_sel' in locals() and school_sel else "school_results.xlsx"
    st.download_button("تنزيل نتائج تحليل المدرسة (Excel)", data=xlsx_bytes,
                       file_name=default_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

combined_dict = {}
combined_dict.update(general_sheets)
combined_dict.update(export_sheets)
if combined_dict:
    combined_name = f"dashboard_all_{school_sel}.xlsx".replace(" ", "_") if 'school_sel' in locals() and school_sel else "dashboard_all.xlsx"
    all_bytes = dfs_to_excel_bytes(combined_dict, images=chart_images, charts_sheet_name="رسوم_بيانية")
    st.download_button("تنزيل ملف موحّد (جداول ورسوم) - Excel", data=all_bytes,
                       file_name=combined_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.caption("يحتوي التنزيل الموحّد على جميع جداول الداشبورد الحالية بالإضافة إلى الرسوم البيانية (مُدرجة كورقة صور داخل ملف Excel).")
